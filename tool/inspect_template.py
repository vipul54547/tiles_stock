"""Inspect a generated stock template: sheets, header text+fill, dropdowns,
freeze panes, and the Lists-sheet legend. Run: python tool/inspect_template.py
"""
import openpyxl

for path in ("build/test_out/tw_template.xlsx", "build/test_out/m_template.xlsx"):
    print("=" * 70)
    print(path)
    wb = openpyxl.load_workbook(path)
    print("  sheets:", wb.sheetnames)
    st = wb["Stock"]
    print("  freeze_panes:", st.freeze_panes)
    print("  --- header row (col: text | fill) ---")
    for c in st[1]:
        if c.value is None:
            continue
        fill = c.fill.fgColor.rgb if c.fill and c.fill.patternType else None
        print(f"    {c.column_letter}: {c.value!r:28} fill={fill}")
    print("  --- data validations (dropdowns) ---")
    for dv in st.data_validations.dataValidation:
        print(f"    {sorted(str(dv.sqref).split())}  ->  {dv.formula1}")
    lists = wb["Lists"]
    print("  --- Lists legend cells (non-vocab text) ---")
    for row in lists.iter_rows():
        for c in row:
            if c.value and any(k in str(c.value) for k in
                               ("guide", "fill", "Purple", "Navy", "Grey")):
                print(f"    {c.coordinate}: {c.value!r}")
print("=" * 70)
