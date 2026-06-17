"""Generate the stock-upload Excel template that matches the importer exactly.

Headers are chosen to hit the importer's header-synonyms so column detection
never fails. Quality / Tile Type / Size get dropdown validation so the stockist
can't mistype a value. Run: python tool/make_stock_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"G:\tiles_stock login details\stock_upload_template.xlsx"

# Allowed values (must match the app)
QUALITIES = ["Premium", "Standard"]
TILE_TYPES = ["PGVT & GVT", "Porcelain", "Ceramic", "Full Body", "DC", "Colour Body"]
SIZES = ["600x1200", "600x600", "400x400", "300x450", "300x600",
         "300x300", "800x1600", "800x1200", "500x500"]
SURFACES = ["None", "P.Glossy", "Matt", "Carving", "High Glossy",
            "Glossy", "Satin Matt", "Rocker", "Sugar", "P.Sugar"]

REQ_FILL = PatternFill("solid", fgColor="1B4F72")   # navy = required
OPT_FILL = PatternFill("solid", fgColor="6C7A89")   # grey = optional
BRAND_FILL = PatternFill("solid", fgColor="6A1B9A")  # purple = brand columns
WHITE = Font(color="FFFFFF", bold=True, size=11)
THIN = Border(*(Side(style="thin", color="D5D8DC"),) * 4)


def style_header(ws, headers, fills):
    for col, (h, fill) in enumerate(zip(headers, fills), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[c.column_letter].width = max(14, len(h) + 3)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def add_dropdown(ws, col_letter, options, n=300):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Pick a value from the list."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{n}")


wb = openpyxl.Workbook()

# ── Sheet 1: Stock List (single brand — use this now) ───────────────────────
ws = wb.active
ws.title = "Stock List"
headers = ["Design Name", "Size", "Quality", "Box Quantity", "Tile Type",
           "Surface", "Box Weight", "Pieces/Box", "Colour"]
fills = [REQ_FILL, REQ_FILL, REQ_FILL, REQ_FILL, REQ_FILL,
         OPT_FILL, OPT_FILL, OPT_FILL, OPT_FILL]
style_header(ws, headers, fills)
# Example rows (replace with your data)
examples = [
    ["Statuario Gold", "800x1600", "Premium", 120, "PGVT & GVT", "P.Glossy", 32, 3, "White"],
    ["Carrara Blue",   "600x1200", "Standard", 60, "Porcelain",  "Matt",     24, 4, "Blue"],
]
for r, row in enumerate(examples, 2):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v).border = THIN
add_dropdown(ws, "B", SIZES)        # Size
add_dropdown(ws, "C", QUALITIES)    # Quality
add_dropdown(ws, "E", TILE_TYPES)   # Tile Type
add_dropdown(ws, "F", SURFACES)     # Surface

# ── Sheet 2: Multi-brand example (for when you add a 2nd brand) ──────────────
ws2 = wb.create_sheet("Multi-brand example")
mheaders = ["Master Design Name", "Size", "Quality", "Box Quantity", "Tile Type",
            "Surface", "Brand-1", "Brand-2"]
mfills = [REQ_FILL, REQ_FILL, REQ_FILL, REQ_FILL, REQ_FILL,
          OPT_FILL, BRAND_FILL, BRAND_FILL]
style_header(ws2, mheaders, mfills)
mex = [
    ["Statuario Gold", "800x1600", "Premium", 120, "PGVT & GVT", "P.Glossy", "Bianco Tera", "Super Terraco"],
    ["Carrara Blue",   "600x1200", "Standard", 60, "Porcelain",  "Matt",     "Azure",       "Blue Wave"],
]
for r, row in enumerate(mex, 2):
    for c, v in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=v).border = THIN
add_dropdown(ws2, "B", SIZES)
add_dropdown(ws2, "C", QUALITIES)
add_dropdown(ws2, "E", TILE_TYPES)
add_dropdown(ws2, "F", SURFACES)

# ── Sheet 3: Instructions ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Instructions")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 70
title = ws3.cell(row=1, column=1, value="How to fill the stock template")
title.font = Font(bold=True, size=14, color="1B4F72")
rows = [
    ("Column", "Need", "Notes"),
    ("Design Name", "REQUIRED", "The tile's name. (Multi-brand: use the Master + brand columns instead.)"),
    ("Size", "REQUIRED", "Must match your sizes: " + ", ".join(SIZES)),
    ("Quality", "REQUIRED", "Premium or Standard (Economy removed)."),
    ("Box Quantity", "REQUIRED", "Number of boxes to add to stock."),
    ("Tile Type", "REQUIRED", "Body type: " + ", ".join(TILE_TYPES)),
    ("Surface", "optional", "Any wording — you map it to a standard finish after upload. Blank = None."),
    ("Box Weight", "optional", "kg per box — used to estimate thickness."),
    ("Pieces/Box", "optional", "Used to compute sq.ft per box."),
    ("Colour", "optional", "Free text."),
    ("", "", ""),
    ("MULTI-BRAND", "", "On the 'Multi-brand example' sheet:"),
    ("Master Design Name", "REQUIRED", "Your internal master name that links the brands."),
    ("<Brand name> cols", "optional", "One column per brand; the HEADER must be the brand's EXACT name in the app. The cell = that tile's name in that brand. The brand you upload to becomes the stock name; ALL brand columns are saved into your Library."),
    ("", "", ""),
    ("Rules", "", "1) Keep the header row. 2) Don't rename required columns. 3) Delete the example rows before importing your data. 4) The first sheet is the one that gets imported."),
]
for r, (a, b, c) in enumerate(rows, 3):
    ca = ws3.cell(row=r, column=1, value=a)
    cb = ws3.cell(row=r, column=2, value=b)
    cc = ws3.cell(row=r, column=3, value=c)
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    if r == 3:
        for cell in (ca, cb, cc):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1B4F72")
    elif b == "REQUIRED":
        cb.font = Font(bold=True, color="1B4F72")
    elif b == "optional":
        cb.font = Font(color="6C7A89")
    else:
        ca.font = Font(bold=True)

wb.save(OUT)
print("Saved:", OUT)
