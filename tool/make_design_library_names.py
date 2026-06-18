"""Design Library — NAMES & LINKING template (the '1 Excel').

This Excel carries ONLY the design identity + per-brand names + cross-brand
linking. One row = one design, linked across every brand the stockist runs.

NOT in this file (set in-app, per design, via the "+" button, mapped to
admin_design_dna via aliases like Surface):
  Design DNA = surface, punch, design joint, glaze, look type, application,
  print type, colour, range  ->  these power buyer SEARCH.

Also NOT here: stock (quantity + quality) = the separate recurring stock upload.

Cell colour:
  RED    fixed identity  (Master ID, Size, Company Design Name)
  ORANGE per-brand names (fill once, optional, add/rename anytime)
Run: python tool/make_design_library_names.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\lenovo\Downloads\design_library_names_template.xlsx"
SIZES = ["600x1200", "600x600", "400x400", "300x450", "300x600",
         "300x300", "800x1600", "800x1200", "500x500"]

RED_H, RED_C       = "C0392B", "F8C9C4"
ORANGE_H, ORANGE_C = "B9770E", "FCE4D6"
WHITE = Font(color="FFFFFF", bold=True, size=11)
THIN  = Border(*(Side(style="thin", color="D5D8DC"),) * 4)

# (header, category, dropdown)  — brand columns: rename header to your real brand
COLS = [
    ("Master ID",            "red",    None),   # blank = new (app assigns)
    ("Size *",               "red",    SIZES),
    ("Company Design Name *","red",    None),   # your internal master name
    ("Brand-1 Design Name",  "orange", None),   # rename header -> brand's real name
    ("Brand-2 Design Name",  "orange", None),   # add one column per extra brand
]
HDR = {"red": (RED_H, RED_C), "orange": (ORANGE_H, ORANGE_C)}


def grid(ws, rows):
    for ci, (h, cat, dd) in enumerate(COLS, 1):
        hh, _ = HDR[cat]
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=hh)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[c.column_letter].width = max(18, len(h) + 2)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    for ri, row in enumerate(rows, 2):
        for ci, (h, cat, dd) in enumerate(COLS, 1):
            _, cc = HDR[cat]
            v = row[ci - 1] if ci - 1 < len(row) else None
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = PatternFill("solid", fgColor=cc)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
    for ci, (h, cat, dd) in enumerate(COLS, 1):
        if dd:
            dv = DataValidation(type="list", formula1='"' + ",".join(dd) + '"',
                                allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(dv)
            le = openpyxl.utils.get_column_letter(ci)
            dv.add(f"{le}2:{le}300")


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Design Library (names)"
# One row = one design, linked across brands. Same name+size = one design.
rows = [
    [None, "800x1600", "Statuario Gold", "Bianco Tera",  "Super Terraco"],
    [None, "600x1200", "Carrara Blue",   "Azure",        "Blue Wave"],
    [None, "600x600",  "Cemento Grey",   "Urban Grey",   ""],   # brand-2 blank = add later
]
grid(ws, rows)

# Legend
g = wb.create_sheet("How it works")
for w, col in [(24, "A"), (66, "B")]:
    g.column_dimensions[col].width = w
g.cell(row=1, column=1, value="Design Library — names & linking").font = \
    Font(bold=True, size=14, color="1B4F72")
items = [
    ("Master ID", "Leave BLANK for a new design — the app assigns a permanent id. That id is the identity, so you can rename the design or any brand name later with NO effect on stock. Put the id back to update an existing design."),
    ("Size", "Required. Same name + different size = a different design (two rows)."),
    ("Company Design Name", "Your own master/internal name for the tile. Required. The single name that links the brand names below."),
    ("Brand-1 / Brand-2 Design Name", "One column PER BRAND you run. Rename the header to the brand's EXACT name in the app. The cell = that tile's name in that brand. Leave blank if you have no name yet — add it later after seeing the design."),
    ("Linking = no duplicates", "Putting all the brand names in ONE row links them as ONE design. If you DON'T link (enter a brand's name as a new row), it just becomes a second design — allowed and harmless; you can merge later."),
    ("NOT in this file: Design DNA", "surface, punch, design joint, glaze, look type, application, print type, colour, range. Set these per design in the app via the + button — type your wording, it maps to the admin master (admin_design_dna) via aliases, exactly like Surface. This DNA powers buyer SEARCH, so fill it well."),
    ("NOT in this file: Stock", "Box quantity + quality go in the separate stock upload (they change every time; the library above is set once)."),
]
r = 3
for k, v in items:
    a = g.cell(row=r, column=1, value=k)
    a.font = Font(bold=True)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = g.cell(row=r, column=2, value=v)
    b.alignment = Alignment(wrap_text=True, vertical="top")
    g.row_dimensions[r].height = 54
    r += 1

wb.save(OUT)
print("Saved:", OUT)
