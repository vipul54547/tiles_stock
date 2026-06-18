"""FINAL Design-Library + Stock template — merges stock_upload_template.xlsx and
master design_library.xlsx, restores the missing Master_Design_ID, and adds the
full attribute set revealed by the library file (Punch, Design Joint, Look Type,
Application, Range).

Cell colour = how often the field changes:
  RED    fixed identity / match key      (Master ID, Brand, Size, Name, Quality)
  GREEN  changes every upload            (Box Quantity)
  BLUE   set first time, then fixed       (Tile Type, Pieces/Box, Box Weight)
  ORANGE fill once, optional, MAPPED      (Surface, Punch, Design Joint, Glaze,
                                           Look Type, Application, Print Type,
                                           Range, Colour) -> matched to an admin
                                           master list via aliases, like Surface.
"*" = required.  Run: python tool/make_design_library_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\lenovo\Downloads\design_library_stock_template.xlsx"

QUALITIES   = ["Premium", "Standard"]
TILE_TYPES  = ["PGVT & GVT", "Porcelain", "Ceramic", "Full Body", "DC", "Colour Body"]
SIZES       = ["600x1200", "600x600", "400x400", "300x450", "300x600",
               "300x300", "800x1600", "800x1200", "500x500"]
# mapped (admin master + stockist alias) lists — samples from your library file
SURFACE     = ["None", "P.Glossy", "Matt", "Carving", "High Glossy", "Glossy",
               "Satin Matt", "Rocker", "Sugar", "P.Sugar"]
PUNCH       = ["None", "Plain", "Texture", "Emboss", "HighDepth"]
DESIGN_JOINT= ["None", "Endless", "Match", "Set"]
GLAZE       = ["None", "Glossy", "Matt", "Sugar", "Metallic"]
LOOK_TYPE   = ["Floral", "Marble", "Wood", "Stone", "Cement/Concrete",
               "Statuario", "Onyx", "Geometric", "Plain/Solid"]
APPLICATION = ["None", "Floor", "Wall", "Floor & Wall", "Outdoor"]
PRINT_TYPE  = ["Full print", "Digital", "Rotary", "Screen", "Nano"]

GREEN_H, GREEN_C   = "2E7D32", "C6EFCE"
RED_H,   RED_C     = "C0392B", "F8C9C4"
BLUE_H,  BLUE_C    = "1B4F72", "BDD7EE"
ORANGE_H, ORANGE_C = "B9770E", "FCE4D6"
WHITE = Font(color="FFFFFF", bold=True, size=11)
THIN  = Border(*(Side(style="thin", color="D5D8DC"),) * 4)
HDR = {"green": (GREEN_H, GREEN_C), "red": (RED_H, RED_C),
       "blue": (BLUE_H, BLUE_C), "orange": (ORANGE_H, ORANGE_C)}

# (header, category, required, dropdown)
COLS = [
    ("Master ID",        "red",    False, None),   # <-- the restored anchor
    ("Brand *",          "red",    True,  None),
    ("Size *",           "red",    True,  SIZES),
    ("Master Design *",  "red",    True,  None),
    ("Quality *",        "red",    True,  QUALITIES),
    ("Box Quantity *",   "green",  True,  None),
    ("Tile Type *",      "blue",   True,  TILE_TYPES),
    ("Pieces/Box",       "blue",   False, None),
    ("Box Weight kg",    "blue",   False, None),
    ("Surface",          "orange", False, SURFACE),
    ("Punch",            "orange", False, PUNCH),
    ("Design Joint",     "orange", False, DESIGN_JOINT),
    ("Glaze",            "orange", False, GLAZE),
    ("Look Type",        "orange", False, LOOK_TYPE),
    ("Application",      "orange", False, APPLICATION),
    ("Print Type",       "orange", False, PRINT_TYPE),
    ("Range",            "orange", False, None),
    ("Colour",           "orange", False, None),
]


def L(i):
    return openpyxl.utils.get_column_letter(i)


def grid(ws, rows):
    for ci, (h, cat, req, dd) in enumerate(COLS, 1):
        hh, _ = HDR[cat]
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=hh)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[c.column_letter].width = max(12, len(h) + 2)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    for ri, row in enumerate(rows, 2):
        for ci, (h, cat, req, dd) in enumerate(COLS, 1):
            _, cc = HDR[cat]
            v = row[ci - 1] if ci - 1 < len(row) else None
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = PatternFill("solid", fgColor=cc)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
    for ci, (h, cat, req, dd) in enumerate(COLS, 1):
        if dd:
            dv = DataValidation(type="list", formula1='"' + ",".join(dd) + '"',
                                allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(dv)
            dv.add(f"{L(ci)}2:{L(ci)}300")


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Design Library + Stock"
# Master ID blank = NEW (system assigns). Row 1 & 4 = SAME tile, different brand
# name -> two designs (allowed, harmless duplicate).
rows = [
    [None, "Brand-1", "800x1600", "Statuario Gold", "Premium", 120, "PGVT & GVT", 3, 32, "P.Glossy", "Plain",   "Endless", "Glossy", "Statuario",       "Floor & Wall", "Full print", "Royal",   "White"],
    [None, "Brand-1", "600x1200", "Carrara Blue",   "Standard", 60, "Porcelain",  4, 24, "Matt",     "Texture", "Endless", "Matt",   "Marble",          "Floor",        "Digital",    "Classic", "Blue"],
    [None, "Brand-1", "600x600",  "Cemento Grey",   "Standard",200, "Ceramic",    6, 18, "Satin Matt","Texture","Endless", "Matt",   "Cement/Concrete", "Wall",         "Digital",    "Urban",   "Grey"],
    [None, "Brand-1", "800x1600", "Statuario White","Premium",  90, "PGVT & GVT", 3, 32, "P.Glossy", "Plain",   "Endless", "Glossy", "Statuario",       "Floor & Wall", "Full print", "Royal",   "White"],
]
grid(ws, rows)

# Sheet 2 — next upload: keep Master ID (or Name) + only the GREEN qty changes.
ws2 = wb.create_sheet("Next upload (re-stock)")
nxt = [
    [None, "Brand-1", "800x1600", "Statuario Gold", "Premium", 75, "", "", "", "", "", "", "", "", "", "", "", ""],
    [None, "Brand-1", "600x1200", "Carrara Blue",   "Standard",40, "", "", "", "", "", "", "", "", "", "", "", ""],
]
grid(ws2, nxt)

# Sheet 3 — legend + guide
g = wb.create_sheet("Legend & Field Guide")
for w, col in [(20, "A"), (16, "B"), (16, "C"), (62, "D")]:
    g.column_dimensions[col].width = w
g.cell(row=1, column=1, value="Cell colour = how often the field changes").font = \
    Font(bold=True, size=14, color="1B4F72")
legend = [
    ("RED",    "red",    "Never change",  "Identity / match key: Master ID, Brand, Size, Master Design, Quality. (Rename is safe IN-APP because the Master ID is the anchor — but in Excel keep these consistent, or fill the Master ID to update a row.)"),
    ("GREEN",  "green",  "Every upload",   "Box Quantity — boxes to ADD this time. The only field you normally re-type."),
    ("BLUE",   "blue",   "First time only","Physical spec set once, then leave blank: Tile Type, Pieces/Box, Box Weight."),
    ("ORANGE", "orange", "Fill once (opt)","Optional, editable, MAPPED to an admin master list via aliases (like Surface today): Surface, Punch, Design Joint, Glaze, Look Type, Application, Print Type, Range, Colour."),
]
r = 3
for c, t in [(1, "Colour"), (2, "Changes"), (4, "Meaning")]:
    g.cell(row=r, column=c, value=t).font = Font(bold=True)
for name, cat, when, desc in legend:
    r += 1
    g.cell(row=r, column=1, value=name).fill = PatternFill("solid", fgColor=HDR[cat][1])
    g.cell(row=r, column=1).font = Font(bold=True)
    g.cell(row=r, column=2, value=when)
    d = g.cell(row=r, column=4, value=desc)
    d.alignment = Alignment(wrap_text=True, vertical="top")
    g.row_dimensions[r].height = 42

r += 2
note = ("KEY POINT — Master ID: leave BLANK for a new design (the app generates a "
        "permanent id). That id — not the name — is the identity, so you can rename "
        "the Master Design or any attribute later with NO effect on stock/boxes. To "
        "update/rename an existing design, put its Master ID back in this column.\n\n"
        "MAPPED attributes: type your OWN wording in any orange column; the app "
        "matches it to the admin master value via aliases (Surface works this way "
        "today). Unknown wording becomes a new alias for admin to confirm.\n\n"
        "DUPLICATES: rows 1 & 4 ('Statuario Gold' / 'Statuario White') are the same "
        "tile under two brand names -> two designs. Allowed and harmless; link/merge "
        "later. Same Name + same Size = one design; same Name + different Size = two.")
nc = g.cell(row=r, column=1, value=note)
nc.alignment = Alignment(wrap_text=True, vertical="top")
g.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
g.row_dimensions[r].height = 150

wb.save(OUT)
print("Saved:", OUT)
