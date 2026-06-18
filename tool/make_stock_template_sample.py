"""Colour-coded stock-upload Excel SAMPLE, coloured by how often each field changes.

Legend (cell colour = change frequency):
  GREEN  – changes EVERY upload            (you re-type it each time)        e.g. Box Quantity
  RED    – FIXED identity, never change     (this is how we find the design)  e.g. Name / Size / Quality
  BLUE   – set the FIRST time, then fixed    (physical spec, leave blank later) e.g. Tile Type / Pcs / Weight
  ORANGE – fill ONCE, optional, editable     (description, mapped to master data) e.g. Surface / Design / Glaze / Print

"*" in a header = REQUIRED. Mapped fields (Surface, Design Type, Glaze Type,
Print Type) work like Surface today: free wording -> matched to an admin master
list via aliases.  Run: python tool/make_stock_template_sample.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\lenovo\Downloads\stock_upload_template_sample.xlsx"

# ── canonical (dropdown) lists already in the app ────────────────────────────
QUALITIES  = ["Premium", "Standard"]
TILE_TYPES = ["PGVT & GVT", "Porcelain", "Ceramic", "Full Body", "DC", "Colour Body"]
SIZES      = ["600x1200", "600x600", "400x400", "300x450", "300x600",
              "300x300", "800x1600", "800x1200", "500x500"]
STOCK_TYPES = ["Continuous", "One Time", "Uncertain"]
# ── mapped (master-data + alias) lists — SAMPLES, admin will manage these ─────
SURFACES    = ["None", "P.Glossy", "Matt", "Carving", "High Glossy", "Glossy",
               "Satin Matt", "Rocker", "Sugar", "P.Sugar"]
DESIGN_TYPES = ["Marble", "Wood", "Stone", "Cement/Concrete", "Statuario",
                "Onyx", "Travertine", "Terrazzo", "Plain/Solid", "Geometric"]
GLAZE_TYPES  = ["Glossy", "Matt", "Sugar", "Carving", "Metallic", "Satin", "Rustic"]
PRINT_TYPES  = ["Digital", "Rotary", "Screen", "Inkjet", "Nano"]

# ── colours ──────────────────────────────────────────────────────────────────
GREEN_H, GREEN_C   = "2E7D32", "C6EFCE"   # changes every upload
RED_H,   RED_C     = "C0392B", "F8C9C4"   # fixed identity
BLUE_H,  BLUE_C    = "1B4F72", "BDD7EE"   # first time then fixed
ORANGE_H, ORANGE_C = "B9770E", "FCE4D6"   # fill once, optional, mapped
WHITE = Font(color="FFFFFF", bold=True, size=11)
THIN  = Border(*(Side(style="thin", color="D5D8DC"),) * 4)

# column -> (header, category, required, dropdown-list-or-None)
COLS = [
    ("Design Name *", "red",    True,  None),
    ("Size *",        "red",    True,  SIZES),
    ("Quality *",     "red",    True,  QUALITIES),
    ("Box Quantity *","green",  True,  None),
    ("Tile Type *",   "blue",   True,  TILE_TYPES),
    ("Pieces/Box",    "blue",   False, None),
    ("Box Weight kg", "blue",   False, None),
    ("Surface",       "orange", False, SURFACES),
    ("Design Type",   "orange", False, DESIGN_TYPES),
    ("Glaze Type",    "orange", False, GLAZE_TYPES),
    ("Print Type",    "orange", False, PRINT_TYPES),
    ("Colour",        "orange", False, None),
    ("Stock Type",    "orange", False, STOCK_TYPES),
]
HDR = {"green": (GREEN_H, GREEN_C), "red": (RED_H, RED_C),
       "blue": (BLUE_H, BLUE_C), "orange": (ORANGE_H, ORANGE_C)}


def col_letter(i):
    return openpyxl.utils.get_column_letter(i)


def write_grid(ws, rows):
    # header
    for ci, (h, cat, req, dd) in enumerate(COLS, 1):
        hh, _ = HDR[cat]
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=hh)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[c.column_letter].width = max(13, len(h) + 2)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    # data, tinted by category
    for ri, row in enumerate(rows, 2):
        for ci, (h, cat, req, dd) in enumerate(COLS, 1):
            _, cc = HDR[cat]
            val = row[ci - 1] if ci - 1 < len(row) else None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=cc)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
    # dropdowns
    for ci, (h, cat, req, dd) in enumerate(COLS, 1):
        if dd:
            dv = DataValidation(type="list", formula1='"' + ",".join(dd) + '"',
                                allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(dv)
            L = col_letter(ci)
            dv.add(f"{L}2:{L}300")


wb = openpyxl.Workbook()

# ── Sheet 1: first upload (full data + a cross-brand DUPLICATE) ───────────────
ws = wb.active
ws.title = "Stock Upload"
first = [
    # Name, Size, Quality, Qty, TileType, Pcs, Wt, Surface, Design, Glaze, Print, Colour, StockType
    ["Statuario Gold", "800x1600", "Premium", 120, "PGVT & GVT", 3, 32, "P.Glossy", "Statuario", "Glossy", "Digital", "White", "Continuous"],
    ["Carrara Blue",   "600x1200", "Standard", 60, "Porcelain",  4, 24, "Matt",     "Marble",    "Matt",   "Rotary",  "Blue",  "One Time"],
    ["Cemento Grey",   "600x600",  "Standard",200, "Ceramic",    6, 18, "Satin Matt","Cement/Concrete","Matt","Digital","Grey", "Uncertain"],
    # DUPLICATE (same physical tile as row 1, different brand name → its own design; harmless)
    ["Statuario White","800x1600", "Premium", 90, "PGVT & GVT", 3, 32, "P.Glossy", "Statuario", "Glossy", "Digital", "White", "Continuous"],
]
write_grid(ws, first)

# ── Sheet 2: NEXT upload — only GREEN changes; once-only fields left blank ────
ws2 = wb.create_sheet("Next upload (re-stock)")
nxt = [
    # RED identity stays IDENTICAL, GREEN qty is new, BLUE/ORANGE blank (already saved)
    ["Statuario Gold", "800x1600", "Premium", 75,  "", "", "", "", "", "", "", "", ""],
    ["Carrara Blue",   "600x1200", "Standard", 40, "", "", "", "", "", "", "", "", ""],
    ["Cemento Grey",   "600x600",  "Standard",150, "", "", "", "", "", "", "", "", ""],
]
write_grid(ws2, nxt)

# ── Sheet 3: legend + field guide ────────────────────────────────────────────
ws3 = wb.create_sheet("Legend & Field Guide")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 64
t = ws3.cell(row=1, column=1, value="Cell colour = how often the field changes")
t.font = Font(bold=True, size=14, color="1B4F72")

legend = [
    ("GREEN",  "green",  "Every upload", "You re-type this each time. Box Quantity (the new stock count)."),
    ("RED",    "red",    "Never change", "Identity / match key. Keep IDENTICAL every upload or it becomes a new design. Name + Size + Quality."),
    ("BLUE",   "blue",   "First time only","Physical spec — set once, then leave blank on later uploads. Tile Type, Pieces/Box, Box Weight."),
    ("ORANGE", "orange", "Fill once (opt)","Optional description, editable anytime, mapped to a master list via aliases. Surface, Design/Glaze/Print Type, Colour, Stock Type."),
]
r = 3
ws3.cell(row=r, column=1, value="Colour").font = Font(bold=True)
ws3.cell(row=r, column=2, value="Changes").font = Font(bold=True)
ws3.cell(row=r, column=4, value="Meaning").font = Font(bold=True)
for name, cat, when, desc in legend:
    r += 1
    hh, cc = HDR[cat]
    a = ws3.cell(row=r, column=1, value=name)
    a.fill = PatternFill("solid", fgColor=cc)
    a.font = Font(bold=True)
    ws3.cell(row=r, column=2, value=when)
    d = ws3.cell(row=r, column=4, value=desc)
    d.alignment = Alignment(wrap_text=True, vertical="top")

r += 2
ws3.cell(row=r, column=1, value="Field").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=r, column=2, value="Need").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=r, column=3, value="Colour").font = Font(bold=True, color="FFFFFF")
ws3.cell(row=r, column=4, value="How it's read / mapped").font = Font(bold=True, color="FFFFFF")
for c in range(1, 5):
    ws3.cell(row=r, column=c).fill = PatternFill("solid", fgColor="1B4F72")
guide = [
    ("Design Name", "REQUIRED", "RED",    "The tile's name in this brand. The match key — keep identical to add stock to the same design."),
    ("Size",        "REQUIRED", "RED",    "Dropdown. Same name + different size = a different design."),
    ("Quality",     "REQUIRED", "RED",    "Dropdown: Premium / Standard."),
    ("Box Quantity","REQUIRED", "GREEN",  "Boxes to ADD this upload. The only field that normally changes each time."),
    ("Tile Type",   "REQUIRED", "BLUE",   "Dropdown (body type). Set once; drives sq.ft + thickness."),
    ("Pieces/Box",  "optional", "BLUE",   "Set once. Used to compute sq.ft per box."),
    ("Box Weight kg","optional","BLUE",   "Set once. Used to estimate thickness."),
    ("Surface",     "optional", "ORANGE", "Free wording -> matched to admin Surface master via aliases (like today). Blank = None."),
    ("Design Type", "optional", "ORANGE", "NEW master table (Marble/Wood/Stone...). Free wording -> matched via aliases."),
    ("Glaze Type",  "optional", "ORANGE", "NEW master table (Glossy/Matt/Sugar...). Free wording -> matched via aliases."),
    ("Print Type",  "optional", "ORANGE", "NEW master table (Digital/Rotary/Screen...). Free wording -> matched via aliases."),
    ("Colour",      "optional", "ORANGE", "Free text."),
    ("Stock Type",  "optional", "ORANGE", "Dropdown: Continuous / One Time / Uncertain (future availability)."),
]
for f, need, colour, how in guide:
    r += 1
    ws3.cell(row=r, column=1, value=f)
    nb = ws3.cell(row=r, column=2, value=need)
    nb.font = Font(bold=True, color="1B4F72") if need == "REQUIRED" else Font(color="6C7A89")
    cat = colour.lower()
    cb = ws3.cell(row=r, column=3, value=colour)
    cb.fill = PatternFill("solid", fgColor=HDR[cat][1])
    hw = ws3.cell(row=r, column=4, value=how)
    hw.alignment = Alignment(wrap_text=True, vertical="top")

r += 2
note = ws3.cell(row=r, column=1,
    value="Duplicates: 'Statuario Gold' and 'Statuario White' (sheet 1) are the SAME tile under two "
          "brand names. They become two designs — that is allowed and harmless (stock stays correct; "
          "you can link/merge them later). On the 'Next upload' sheet only the GREEN box quantity is "
          "filled — everything else is already saved.")
note.alignment = Alignment(wrap_text=True, vertical="top")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws3.row_dimensions[r].height = 60

wb.save(OUT)
print("Saved:", OUT)
